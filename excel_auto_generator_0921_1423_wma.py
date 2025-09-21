# 代码生成时间: 2025-09-21 14:23:50
import os
from sanic import Sanic
from sanic.response import json, file
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# 初始化Sanic应用
app = Sanic('ExcelAutoGenerator')

# 定义一个生成Excel表格的函数
def generate_excel(data: list, filename: str) -> str:
    # 创建工作簿
    wb = Workbook()
    # 激活第一个工作表
    ws = wb.active
    # 将数据写入工作表
    for row in data:
        ws.append(row)
    # 保存工作簿
    wb.save(filename)
    return filename

# 定义一个加载Excel表格的函数
def load_excel(filename: str) -> list:
    try:
        # 加载工作簿
        wb = load_workbook(filename)
        # 获取第一个工作表
        ws = wb.active
        # 读取所有行数据
        data = [[cell.value for cell in row] for row in ws.iter_rows()]
        # 删除标题行（如果有）
        if data and data[0] == ['标题']:
            data = data[1:]
        return data
    except InvalidFileException:
        raise ValueError('无效的Excel文件')

# 定义一个HTTP路由，用于生成Excel表格
@app.route('/create_excel', methods=['POST'])
async def create_excel(request):
    # 获取请求数据
    data = request.json.get('data', [])
    filename = request.json.get('filename', 'generated_excel.xlsx')
    try:
        # 生成Excel表格
        generate_excel(data, filename)
        # 返回成功响应
        return json({'message': 'Excel表格生成成功', 'filename': filename})
    except Exception as e:
        # 返回错误响应
        return json({'error': str(e)})

# 定义一个HTTP路由，用于加载Excel表格
@app.route('/load_excel', methods=['POST'])
async def load_excel(request):
    # 获取请求数据
    filename = request.json.get('filename', 'generated_excel.xlsx')
    try:
        # 加载Excel表格
        data = load_excel(filename)
        # 返回成功响应
        return json({'data': data})
    except Exception as e:
        # 返回错误响应
        return json({'error': str(e)})

# 定义一个HTTP路由，用于下载Excel表格
@app.route('/download_excel/<filename:path>', methods=['GET'])
async def download_excel(request, filename):
    # 检查文件是否存在
    file_path = f'{os.getcwd()}/{filename}'
    if not os.path.isfile(file_path):
        return json({'error': '文件不存在'})
    # 返回文件下载响应
    return file(file_path, filename=filename)

# 运行Sanic应用
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000)